from flask import Flask, render_template, Response,make_response,send_from_directory
from flask_caching import Cache
import cv2
import numpy as np
import json,time
from time import time,sleep
import PoseModule as pm
import openpyxl as opxl
import input as inp

from multiprocessing import Process
from time import sleep

def f(time):
    sleep(time)


def run_with_limited_time(func, args, kwargs, time):
    p = Process(target=func, args=args, kwargs=kwargs)
    p.start()
    p.join(time)
    if p.is_alive():
        p.terminate()
        return False

    return True

app = Flask(__name__)
filepath="data.xlsx"

wb = opxl.load_workbook(filepath)
sheet = wb.active
def rdmat(sheet,min_row,min_col,max_row,max_col):
    rows_iter = sheet.iter_cols(min_row,min_col,max_row,max_col)
    listt = [[cell.value for cell in list(row)] for row in rows_iter][0]
    return listt

def wtmat(sheet,datalist):
    # print(datalist)
    row=len(datalist)
    col=len(datalist[0])
    for i in range(row):
        for j in range(col):
            sheet.cell((5+i), (2+j)).value = datalist[i][j]
    return sheet
def wtmat2(sheet,datalist):
    # print(datalist)
    row=len(datalist)
    col=len(datalist[0])
    for i in range(row):
        for j in range(col):
            sheet.cell((5+i), (5+j)).value = datalist[i][j]
    return sheet
def wrt_data2(datalist):
    global  sheet
    global  wb
    sheet=wtmat2(sheet, datalist)
    wb.save(filepath)

    return  0
def wrt_data(datalist):
    global  sheet
    global  wb
    sheet=wtmat(sheet, datalist)
    wb.save(filepath)

    return  0

def get_data(min_row, min_col, max_row, max_col):
    wb = opxl.load_workbook(filepath)
    sheet = wb.active
    dd=rdmat(sheet, min_row, min_col, max_row, max_col)
    # wb.close()
    return dd
list_data =get_data(9,10,5,30)
list_data_out =[]
list_data_out2 =[]
i_l_d=0
#camera = cv2.VideoCapture('rtsp://freja.hiof.no:1935/rtplive/_definst_/hessdalen03.stream')  # use 0 for web camera
@app.route('/<path:path>')
def static_file(path):
    return app.send_static_file(path)
# @app.route('/<path:path>')
# def send_js(path):
#     return send_from_directory('static', path)
# camera = cv2.VideoCapture(0)  # use 0 for web camera
camera = cv2.VideoCapture(inp.VideoCapture01)  # use 0 for web camera
camera2 = cv2.VideoCapture(inp.VideoCapture02)  # use 0 for web camera
detector = pm.poseDetector()
global y_data
y_data=0
global y_data2
y_data2=0
#  for cctv camera use rtsp://username:password@ip_address:554/user=username_password='password'_channel=channel_number_stream=0.sdp' instead of camera
# for local webcam use cv2.VideoCapture(0)
def tran(success,img,count,dir,pTime):
        img = cv2.resize(img, (1280, 720))
        # img = cv2.imread("AiTrainer/test.jpg")
        img = detector.findPose(img, False)
        lmList = detector.findPosition(img, False)
        # print(lmList)
        if len(lmList) != 0:
            # Right Arm
            # angle = detector.findAngle(img, 23, 25, 27)

            angle = detector.findAngle(img, inp.findAngle[0], inp.findAngle[1], inp.findAngle[2])
            # # Left Arm
            # angle = detector.findAngle(img, 11, 13, 15,False)
            per = np.interp(angle, (210, 310), (0, 100))
            bar = np.interp(angle, (220, 310), (650, 100))
            # print(angle, per)

            # Check for the dumbbell curls
            color = (255, 0, 255)
            global y_data
            count=angle
            y_data= angle
            # if per == 100:
            #     color = (0, 255, 0)
            #     if dir == 0:
            #         count += 0.5
            #         dir = 1
            # if per == 0:
            #     color = (0, 255, 0)
            #     if dir == 1:
            #         count += 0.5
            #         dir = 0
            # print(count)

            # Draw Bar
            # cv2.rectangle(img, (1100, 100), (1175, 650), color, 3)
            # cv2.rectangle(img, (1100, int(bar)), (1175, 650), color, cv2.FILLED)
            # cv2.putText(img, f'{int(per)} %', (1100, 75), cv2.FONT_HERSHEY_PLAIN, 4,
            #             color, 4)

            # Draw Curl Count
            # cv2.rectangle(img, (0, 450), (250, 720), (0, 255, 0), cv2.FILLED)
            cv2.putText(img, str(int(count)), (45, 670), cv2.FONT_HERSHEY_PLAIN, 15,
                        (255, 0, 0), 25)

        cTime = time()
        fps = 1 / (cTime - pTime)
        pTime = cTime
        # cv2.putText(img, str(int(fps)), (50, 100), cv2.FONT_HERSHEY_PLAIN, 5,
        #             (255, 0, 0), 5)
        return success,img,count,dir,pTime


def tran2(success, img, count, dir, pTime):
    img = cv2.resize(img, (1280, 720))
    # img = cv2.imread("AiTrainer/test.jpg")
    img = detector.findPose(img, False)
    lmList = detector.findPosition(img, False)
    # print(lmList)
    if len(lmList) != 0:
        # Right Arm
        # angle = detector.findAngle(img, 23, 25, 27)

        angle = detector.findAngle(img, inp.findAngle[0], inp.findAngle[1], inp.findAngle[2])
        # # Left Arm
        # angle = detector.findAngle(img, 11, 13, 15,False)
        per = np.interp(angle, (210, 310), (0, 100))
        bar = np.interp(angle, (220, 310), (650, 100))
        # print(angle, per)

        # Check for the dumbbell curls
        color = (255, 0, 255)
        global y_data2
        count = angle
        y_data2 = angle
        # if per == 100:
        #     color = (0, 255, 0)
        #     if dir == 0:
        #         count += 0.5
        #         dir = 1
        # if per == 0:
        #     color = (0, 255, 0)
        #     if dir == 1:
        #         count += 0.5
        #         dir = 0
        # print(count)

        # Draw Bar
        # cv2.rectangle(img, (1100, 100), (1175, 650), color, 3)
        # cv2.rectangle(img, (1100, int(bar)), (1175, 650), color, cv2.FILLED)
        # cv2.putText(img, f'{int(per)} %', (1100, 75), cv2.FONT_HERSHEY_PLAIN, 4,
        #             color, 4)

        # Draw Curl Count
        # cv2.rectangle(img, (0, 450), (250, 720), (0, 255, 0), cv2.FILLED)
        cv2.putText(img, str(int(count)), (45, 670), cv2.FONT_HERSHEY_PLAIN, 15,
                    (255, 0, 0), 25)

    cTime = time()
    fps = 1 / (cTime - pTime)
    pTime = cTime
    # cv2.putText(img, str(int(fps)), (50, 100), cv2.FONT_HERSHEY_PLAIN, 5,
    #             (255, 0, 0), 5)
    return success, img, count, dir, pTime


def gen_frames(camera):  # generate frame by frame from camera
    while True:
            count = 0
            dir = 0
            pTime = 0
            # Capture frame-by-frame
            success, frame = camera.read()  # read the camera frame


            try:
                if inp.dedect01:
                    success, frame, count, dir, pTime = tran(success, frame, count, dir, pTime)

                ret, buffer = cv2.imencode('.jpg', frame)
                frame = buffer.tobytes()
                sleep(inp.time_sleep)
                yield (b'--frame\r\n'
                       b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')  # concat frame one by one and show result

            except:
                pass

def gen_frames2(camera2):  # generate frame by frame from camera
    while True:


            count2 = 0
            dir2 = 0
            pTime2 = 0
            # Capture frame-by-frame
            success2, frame2 = camera2.read()  # read the camera frame
            try:
                # success2, frame2, count2, dir2, pTime2 = tran2(success2, frame2, count2, dir2, pTime2)
                if inp.dedect02:
                    success2, frame2, count2, dir2, pTime2 = tran2(success2, frame2, count2, dir2, pTime2)

                ret2, buffer2 = cv2.imencode('.jpg', frame2)
                frame2 = buffer2.tobytes()

                sleep(inp.time_sleep)
                yield (b'--frame2\r\n'
                       b'Content-Type: image/jpeg\r\n\r\n' + frame2 + b'\r\n')  # concat frame one by one and show result
            except:
                pass




@app.route('/video_feed')
def video_feed():
    return Response(gen_frames(camera), mimetype='multipart/x-mixed-replace; boundary=frame')
@app.route('/video_feed2')
def video_feed2():
    #Video streaming route. Put this in the src attribute of an img tag

    return Response(gen_frames2(camera2), mimetype='multipart/x-mixed-replace; boundary=frame2')


@app.route('/mal')
def index2():

    global  camera
    camera = cv2.VideoCapture(inp.VideoCapture01)  # use 0 for web camera
    global camera2
      # use 0 for web camera
    camera2 = cv2.VideoCapture(inp.VideoCapture02)

    """Video streaming home page."""
    return render_template('index2.html')

@app.route('/')
def main():

    global  camera
    camera = cv2.VideoCapture(inp.VideoCapture01)  # use 0 for web camera
    global camera2
      # use 0 for web camera
    camera2 = cv2.VideoCapture(inp.VideoCapture02)

    """Video streaming home page."""
    return render_template('main.html')
@app.route('/dr')
def index():

    global  camera
    camera = cv2.VideoCapture(inp.VideoCapture01)  # use 0 for web camera
    global camera2
      # use 0 for web camera
    camera2 = cv2.VideoCapture(inp.VideoCapture02)

    """Video streaming home page."""
    return render_template('index.html')
@app.route('/aa', methods=["GET", "POST"])
def my_link():
    # print("ddd")
    wrt_data(list_data_out)

    wrt_data2(list_data_out2)
    response = make_response(json.dumps("data"))
    response.content_type = 'application/json'
    return response

@app.route('/data', methods=["GET", "POST"])
def data():
        # run_with_limited_time(f, (0.5,), {}, 2.5)
    # run_with_limited_time(f, (3.5,), {}, 2.5)
        global i_l_d
        if i_l_d == len(list_data) - 1:
            i_l_d = 0
        else:
            i_l_d += 1
        global list_data_out
        global list_data_out2

        if 1 == 1:
            list_data_out.append([time() * 1000, y_data])
            list_data_out2.append([time() * 1000, y_data2])
            data = [[time() * 1000, y_data], [time() * 1000, list_data[i_l_d]], [time() * 1000, y_data2], ]
            response = make_response(json.dumps(data))
            response.content_type = 'application/json'
            return response
        #
        # data = [0, 0, 0, ]
        # response = make_response(json.dumps(data))
        # response.content_type = 'application/json'
        # return response






if __name__ == '__main__':
    config = {
        "DEBUG": True,  # some Flask specific configs
        # "CACHE_TYPE": "SimpleCache",  # Flask-Caching related configs
        # "CACHE_DEFAULT_TIMEOUT": 20
    }

    # app.config.from_mapping(config)
    # cache = Cache(app)
    # print(cache)

    # while True:
    #     try:
    app.run(debug=True,threaded=True)
        # except:
        #     pass
        # else:
        #     pass


